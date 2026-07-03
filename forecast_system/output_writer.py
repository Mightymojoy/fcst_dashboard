# -*- coding: utf-8 -*-
"""output_writer.py — v5.0 公式写入"""
import os, shutil
from datetime import datetime
from openpyxl import load_workbook, utils
from config import *

WEIGHTS = [0.05, 0.10, 0.15, 0.20, 0.25, 0.30]
PRED_COLS = {'7':'I','8':'J','9':'K','10':'L','11':'M'}
SEAS = {1:0.9475,2:0.4193,3:0.5491,4:0.5309,5:1.1572,6:1.3598,7:0.9512,8:1.3004,9:1.078,10:1.6053,11:1.2662,12:0.8352}

def backup_template(p):
    b = p.replace('.xlsx',f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    shutil.copy2(p,b); print(f'[备份] {b}')

def write_predictions(template_path, predict_sheet, predictions):
    print('[写入] 写入Excel公式...')
    wb = load_workbook(template_path)

    # 建参数sheet
    if '计算参数' in wb.sheetnames:
        del wb['计算参数']
    ps = wb.create_sheet('计算参数')
    ps['A1']='月份'; ps['B1']='季节系数'
    for m in range(1,13):
        ps.cell(row=m+1,column=1,value=f'{m}月')
        ps.cell(row=m+1,column=2,value=SEAS[m])

    # 读实际sheet建立SKU→行号索引
    ws_act = wb[ACTUAL_SHEET]
    sku_to_row = {}
    for r in range(2, 200):
        sku = str(ws_act.cell(row=r,column=1).value or '').strip()
        if sku:
            sku_to_row[sku] = r

    # 停产品集合
    disco = set(DISCONTINUED_SKUS)
    new_set = set(NEW_JUNE15_SKUS.keys())
    upgrade_set = set(UPGRADE_NEW_SKUS)
    old_set = set(OLD_DISCONTINUING_SKUS)

    ws_pred = wb[predict_sheet]
    param_row = 14  # 参数表从14行开始写趋势因子

    ps.cell(row=13,column=4,value='SKU')
    ps.cell(row=13,column=5,value='趋势因子')
    ps.cell(row=13,column=6,value='月均值')
    ps.cell(row=13,column=7,value='7-11合计')

    written = 0
    for r in range(2, 200):
        sku = str(ws_pred.cell(row=r,column=1).value or '').strip()
        if not sku:
            continue
        if sku not in predictions:
            continue

        # 找实际数据行
        act_row = sku_to_row.get(sku)
        if not act_row:
            continue

        # 读1-6月实际
        vals = []
        for c in range(3, 9):  # C=3(1月) ... H=8(6月)
            v = ws_act.cell(row=act_row,column=c).value
            if v and str(v).strip() and float(v) > 0:
                vals.append(float(v))

        n = len(vals)
        p = predictions[sku]

        if sku in disco:
            for pk in ['7','8','9','10','11']:
                ws_pred[f'{PRED_COLS[pk]}{r}'] = 0
            written += 1
            continue

        if sku in new_set or sku in upgrade_set:
            for pk in ['7','8','9','10','11']:
                v = p.get(f'2026-{pk}',0)
                ws_pred[f'{PRED_COLS[pk]}{r}'] = int(v) if v else 0
            written += 1
            continue

        if sku in old_set:
            for pk in ['7','8','9','10','11']:
                mn = int(pk)
                if mn >= 10:
                    ws_pred[f'{PRED_COLS[pk]}{r}'] = 0
                else:
                    v = p.get(f'2026-{pk}',0)
                    ws_pred[f'{PRED_COLS[pk]}{r}'] = int(v) if v else 0
            written += 1
            continue

        # ===== 正常SKU：写公式 =====
        if n <= 1:
            avg = vals[0] if n==1 else 0
            for pk in ['7','8','9','10','11']:
                ws_pred[f'{PRED_COLS[pk]}{r}'] = int(round(avg))
            written += 1
            continue

        # 趋势因子的计算
        if n >= 4:
            r3 = sum(vals[-3:])/3
            p3 = sum(vals[-min(3,n-3):-3])/min(3,n-3) if n>=6 else sum(vals[:2])/2
            trend = (r3-p3)/p3 if p3>0 else 0
            trend = max(-0.3, min(0.3, trend))
        else:
            trend = 0

        # 趋势因子的Excel公式
        # =((近3月均)-(前3月均))/(前3月均)
        # 用实际数据行号
        def ref(col):
            return f"'{ACTUAL_SHEET}'!{utils.get_column_letter(col)}{act_row}"

        # 构建加权均值公式项
        w_parts = []
        w_total = 0
        for i in range(max(0,6-n), 6):
            ci = i + 3  # C=3...H=8
            wt = WEIGHTS[i]
            w_total += wt
            w_parts.append(f"{ref(ci)}*{wt}")

        w_formula = '+'.join(w_parts)

        # 趋势公式（在参数表中已计算，直接引用）
        ps.cell(row=param_row,column=4,value=sku)
        ps.cell(row=param_row,column=5,value=round(trend,4))
        ps.cell(row=param_row,column=6,value=round(sum(vals)/n,1))

        pred_sum = sum(p.values())
        ps.cell(row=param_row,column=7,value=pred_sum)

        # 写入预测公式
        for pk in ['7','8','9','10','11']:
            mn = int(pk)
            col = PRED_COLS[pk]
            sea_ref = f'计算参数!$B${mn+1}'
            trend_ref = f'计算参数!$E${param_row}'
            formula = f'=ROUND(({w_formula})/{w_total}*{sea_ref}*(1+{trend_ref}),0)'
            ws_pred[f'{col}{r}'] = formula

        param_row += 1
        written += 1

    wb.save(template_path)
    print(f'[写入] 完成 {written} 个SKU')
    print(f'[写入] Excel公式可追溯: 实际数据→加权均值×季节×趋势')
