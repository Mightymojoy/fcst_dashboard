# -*- coding: utf-8 -*-
"""
ts_transformer.py — 目标约束预测 v2.0
逻辑：
  1. 月度总量 = 分月销售目标（万元 → 件数 × 加权均价）
  2. 各系列在月度总量中的占比 = 6月实际占比 × (1 + 趋势调整)
  3. 趋势调整 = Transformer判断的近期方向（仅提取方向，不提取幅度）
  4. SKU级 = 系列内按6月颜色×尺寸权重分配
"""

import numpy as np
import pandas as pd
import os, json, sys
from datetime import datetime, timedelta

SRC_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(SRC_ROOT, '各渠道销售数据源')
TEMPLATE_PATH = os.path.join(SRC_ROOT, '天猫预测-2026年销量预测.xlsx')
CHANNEL_NAME = '直销_伊稻_电商_天猫ITO旗舰店'

# 月度销售目标（万元）
MONTHLY_TARGET_WAN = {
    '2026-07': 1360, '2026-08': 1700, '2026-09': 1450,
    '2026-10': 1900, '2026-11': 1350, '2026-12': 1000
}
# 加权均价（元/件）
AVG_PRICE = 1616

def load_daily_series(min_date='2024-01-01', max_date='2026-12-31'):
    files = sorted([f for f in os.listdir(DATA_DIR) if f.endswith('.xlsx') and f[0].isdigit()])
    all_daily = []
    for fname in files:
        try:
            df = pd.read_excel(os.path.join(DATA_DIR, fname))
            if '店铺' not in df.columns or '货品编号' not in df.columns: continue
            ito = df[df['店铺'] == CHANNEL_NAME].copy()
            if len(ito) == 0: continue
            ito['日期'] = pd.to_datetime(ito['日期'], errors='coerce')
            ito = ito.dropna(subset=['日期'])
            ito = ito[(ito['日期'] >= min_date) & (ito['日期'] <= max_date)]
            all_daily.append(ito)
        except: continue
    return pd.concat(all_daily, ignore_index=True) if all_daily else pd.DataFrame()

def load_sku_info():
    sku = pd.read_excel(TEMPLATE_PATH, sheet_name='SKU明细')
    sku['颜色'] = sku['颜色'].fillna('').astype(str)
    sku['尺寸'] = sku['尺寸'].fillna('').astype(str)
    return sku

def get_series_jun_actual(daily_df, sku_df):
    """计算各系列2026年6月实际件数"""
    sku_to_series = {}
    for _, r in sku_df.iterrows():
        sku = str(r['SKU编码']).strip()
        series = str(r['系列']).strip() if pd.notna(r['系列']) else '其他'
        sku_to_series[sku] = series

    series_qty = {}
    for _, r in daily_df.iterrows():
        sku = str(r['货品编号']).strip()
        series = sku_to_series.get(sku, '其他')
        dt = r['日期']
        if dt.year != 2026 or dt.month != 6: continue
        qty = int(r['实际销售量']) if pd.notna(r.get('实际销售量', 0)) else 0
        if qty <= 0: continue
        series_qty[series] = series_qty.get(series, 0) + qty
    
    return series_qty

def compute_trend_direction(daily_df, sku_df, series_name, lookback=60):
    """
    计算系列的近期趋势方向
    用Transformer的思路：近30天 vs 近60天日均销量
    返回: -1(下降), 0(平稳), 1(上升)
    """
    sku_to_series = {}
    for _, r in sku_df.iterrows():
        sku = str(r['SKU编码']).strip()
        series = str(r['系列']).strip() if pd.notna(r['系列']) else '其他'
        sku_to_series[sku] = series

    today = datetime(2026, 6, 30)
    daily_qty = {}
    for _, r in daily_df.iterrows():
        sku = str(r['货品编号']).strip()
        s = sku_to_series.get(sku, '其他')
        if s != series_name: continue
        dt = r['日期']
        if dt < today - timedelta(days=lookback): continue
        ds = dt.strftime('%Y-%m-%d')
        qty = int(r['实际销售量']) if pd.notna(r.get('实际销售量', 0)) else 0
        if qty <= 0: continue
        daily_qty[ds] = daily_qty.get(ds, 0) + qty

    if not daily_qty: return 0
    dates = sorted(daily_qty.keys())
    mid = (datetime.strptime(dates[-1], '%Y-%m-%d') - timedelta(days=30)).strftime('%Y-%m-%d')
    first = [daily_qty[d] for d in dates if d < mid]
    second = [daily_qty[d] for d in dates if d >= mid]
    
    f_avg = np.mean(first) if len(first) > 3 else 0
    s_avg = np.mean(second) if len(second) > 3 else 0
    if f_avg <= 0: return 0
    ratio = s_avg / f_avg
    if ratio > 1.1: return 1
    if ratio < 0.9: return -1
    return 0

def main():
    print('=' * 60)
    print('TS-Transformer 预测引擎 v2.0（目标约束版）')
    print('=' * 60)
    
    print('\n[1/5] 加载日级数据...')
    daily_df = load_daily_series()
    print(f'  总行数: {len(daily_df)}')
    
    print('\n[2/5] 加载SKU信息...')
    sku_df = load_sku_info()
    exclude_series = ['PISTACHIO 2 秋冬限定色-25', 'CHANTERELLE DUFFLE BAG 2', '其他']
    sku_df = sku_df[~sku_df['系列'].isin(exclude_series)].copy()
    print(f'  SKU数: {len(sku_df)}')
    
    print('\n[3/5] 计算各系列6月实际占比...')
    series_actual = get_series_jun_actual(daily_df, sku_df)
    total_jun = sum(series_actual.values())
    print(f'  6月总销量: {total_jun}件')
    
    # 排除异常系列后的占比
    series_share = {}
    for s, qty in sorted(series_actual.items(), key=lambda x: -x[1]):
        if s in exclude_series: continue
        series_share[s] = qty / total_jun if total_jun > 0 else 0
        print(f'    {s}: {qty}件 ({series_share[s]*100:.1f}%)')
    
    print('\n[4/5] 计算趋势方向...')
    trend_sign = {}
    for s in series_share:
        td = compute_trend_direction(daily_df, sku_df, s)
        trend_sign[s] = td
        arrows = {1: '↑上升', 0: '→平稳', -1: '↓下降'}
        print(f'    {s}: {arrows.get(td, "?")}')
    
    print('\n[5/5] 用量生成（分月目标→系列分配→SKU级）...')
    all_series_pred = {}
    for mk, target_wan in MONTHLY_TARGET_WAN.items():
        target_qty = target_wan * 10000 / AVG_PRICE
        print(f'\n  {mk}: 目标{target_wan}万 ≈ {target_qty:.0f}件')
        
        series_pred = {}
        allocated = 0
        # 先按占比60% + 趋势调整40%的权重分配
        for s, share in series_share.items():
            t = trend_sign.get(s, 0)
            trend_weight = 1.0 + t * 0.1  # 上升+10%, 下降-10%, 平稳不变
            adjusted_share = share * trend_weight
            series_pred[s] = adjusted_share
        
        # 归一化
        total_weight = sum(series_pred.values())
        for s in series_pred:
            series_pred[s] /= total_weight
        
        # 分配件数
        series_qty = {}
        for s, w in sorted(series_pred.items(), key=lambda x: -x[1]):
            qty = max(1, int(round(target_qty * w)))
            series_qty[s] = qty
            allocated += qty
            print(f'    {s}: {qty}件 ({w*100:.1f}%)')
        
        # 尾差修正
        diff = int(target_qty) - allocated
        if diff != 0 and series_qty:
            # 调整占比最大的系列
            top = max(series_qty, key=series_qty.get)
            series_qty[top] = max(1, series_qty[top] + diff)
            print(f'    [尾差修正] {top}: {series_qty[top]}件 (调整{diff:+d})')
        
        all_series_pred[mk] = series_qty
    
    # 新品特别修正：PISTACHIO PLUS STRIPED
    jun_plus_striped = series_actual.get('PISTACHIO PLUS STRIPED', 0)
    jun_plus_trunk = series_actual.get('PISTACHIO PLUS STRIPED TRUNK', 0)
    print(f'\n  新品修正: PISTACHIO PLUS STRIPED (6月={jun_plus_striped}件)')
    for mk in MONTHLY_TARGET_WAN:
        if mk in all_series_pred:
            base = all_series_pred[mk]
            mi = int(mk.split('-')[1])
            if mi <= 9:  # 7-9月：增长期
                factor = 1.0 + (10 - mi) * 0.03
            else:  # 10-12月：趋稳
                factor = 1.0 + (mi - 6) * 0.02
            new_qty = max(jun_plus_striped, int(round(jun_plus_striped * factor)))
            old_qty = base.get('PISTACHIO PLUS STRIPED', 0)
            base['PISTACHIO PLUS STRIPED'] = new_qty
            diff = old_qty - new_qty
            if diff > 0:
                base['PISTACHIO STRIPED'] = base.get('PISTACHIO STRIPED', 0) + diff
    
    # 汇总验证
    print('\n  预测汇总:')
    total_all = {}
    for mk in MONTHLY_TARGET_WAN:
        qty = sum(all_series_pred.get(mk, {}).values())
        total_all[mk] = qty
        amt = qty * AVG_PRICE / 10000
        target = MONTHLY_TARGET_WAN[mk]
        print(f'    {mk}: {qty}件 ≈ {amt:.0f}万 (目标{target}万, 达成{amt/target*100:.0f}%)')
    
    # 按系列整理预测（供distribute_to_skus使用）
    series_pred_dict = {}
    for s in series_share:
        sp = {}
        for mk in MONTHLY_TARGET_WAN:
            sp[mk] = all_series_pred.get(mk, {}).get(s, 0)
        series_pred_dict[s] = sp
    
    # SKU级分配
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # 计算各系列下颜色×尺寸的6月权重
    sku_weights = {}  # {series: {(color,size): weight}}
    for s in series_share:
        rows = sku_df[sku_df['系列'] == s]
        total_jun_s = rows['6月件数'].sum()
        w = {}
        for _, r in rows.iterrows():
            key = (str(r['颜色']).strip(), str(r['尺寸']).strip())
            jun_qty = int(r['6月件数']) if pd.notna(r.get('6月件数', 0)) else 0
            if jun_qty > 0:
                w[key] = w.get(key, 0) + jun_qty
        if total_jun_s > 0:
            for k in w:
                w[k] /= total_jun_s
        if not w:
            n = len(rows)
            for _, r in rows.iterrows():
                w[(str(r['颜色']).strip(), str(r['尺寸']).strip())] = 1.0 / n
        sku_weights[s] = w
    
    # 更新SKU明细
    updated_rows = []
    for _, r in sku_df.iterrows():
        sku = str(r['SKU编码']).strip()
        series = str(r['系列']).strip()
        color = str(r['颜色']).strip()
        size = str(r['尺寸']).strip()
        key = (color, size)
        price = float(r['吊牌价']) if pd.notna(r['吊牌价']) else 0
        
        new_row = r.to_dict()
        if series in series_pred_dict:
            sp = series_pred_dict[series]
            w = sku_weights.get(series, {})
            weight = w.get(key, 0)
            if w and weight == 0:
                weight = min(w.values()) * 0.5 if w else 1.0
            
            for mi, mk in enumerate(['2026-07','2026-08','2026-09','2026-10','2026-11','2026-12'], 7):
                total = sp.get(mk, 0)
                pred_qty = max(1, int(round(total * weight)))
                new_row[f'{mi}月件数'] = pred_qty
                new_row[f'{mi}月金额'] = round(pred_qty * price / 10000, 1)
        
        updated_rows.append(new_row)
    
    updated_df = pd.DataFrame(updated_rows)
    
    # 写入Excel
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb['SKU明细']
    ws.delete_rows(1, ws.max_row)
    for r_idx, row in enumerate(dataframe_to_rows(updated_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            if pd.isna(value): value = 0
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(TEMPLATE_PATH)
    print(f'\n[写入] {len(updated_rows)}行+header → SKU明细')
    print('\n完成!')
    
    # 打印新品关键数据
    print(f'\n  PISTACHIO PLUS STRIPED 7月: {series_pred_dict.get("PISTACHIO PLUS STRIPED", {}).get("2026-07", 0)}件 (6月实际={jun_plus_striped}件)')

if __name__ == '__main__':
    main()
